import csv
from openpyxl import load_workbook
'''
解析数据文件
'''

def do_csv(csvpath):
    data = []
    with open(csvpath,mode='r',encoding='utf8') as f:
        reader = csv.reader(f)
        next(reader)
        for line in reader:
            print(line)
            data.append(tuple(line))

    return data

def do_xls(xlspath):
    workbook = load_workbook(filename=xlspath)
    #sheetname sheet的名
    sheet_name = workbook.sheetnames
    #当前打开的s
    worksheet = workbook['测试1']

    #整合xlsx数据  方法1
    # data = []
    # testdata= []
    # for i in range(1,4):
    #     for x in range(1,4):
    #         data.append(worksheet.cell(row=i,column=x).value)
    #     testdata.append(tuple(data))
    #整合xlsx数据  方法2
    testdata=[]
    for row in worksheet.iter_rows(min_col=0,max_col=3,min_row=0,max_row=3,values_only=True):
        # print(row)
        testdata.append(row)

    return testdata

if __name__ == '__main__':
    #print(do_csv('../data/data.csv'))
    print(do_xls('../data/test.xlsx'))
